import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import threading
import tkinter as tk
from tkinter import ttk
import os

# Função para criar a janela de carregamento
def animacao():
    global loanding, progress, progress_label, start_button
    loanding = tk.Tk()
    loanding.title("Processamento de Query")
    loanding.iconbitmap("melhoria.ico")
    loanding.geometry("400x200")
    loanding.resizable(False, False)
    loanding.configure(bg="#191919")

    # Rótulo
    title_label = tk.Label(loanding, text="Processamento de Query", font=("Arial", 14, "bold"), fg="white", bg="#191919")
    title_label.pack(pady=10)

    # Rótulo com a porcentagem
    progress_label = tk.Label(loanding, text="Aguardando início...", font=("Arial", 12), fg="white", bg="#191919")
    progress_label.pack(pady=5)

    # Frame para a barra de progresso
    frame = tk.Frame(loanding, bg="#1C1C1C", padx=3, pady=3)
    frame.pack(pady=10)

    # Barra de progresso
    progress = ttk.Progressbar(frame, orient="horizontal", length=300, mode="determinate")
    progress.pack(pady=5)
    progress["value"] = 0

    # Botão para iniciar o processo
    start_button = tk.Button(loanding, text="Iniciar", command=start_excel_generation, font=("Arial", 12), bg="#4CAF50", fg="white")
    start_button.pack(pady=10)

    loanding.update_idletasks()

# Atualizar a barra de progresso e o título
def update_progresso(value, title):
    progress["value"] = value
    progress_label.config(text=f"{value}% concluído")
    loanding.title(title)
    loanding.update_idletasks()

# Função para mostrar mensagem de conclusão
def concluido():
    for widget in loanding.winfo_children():
        widget.destroy()  # Remove todos os widgets anteriores

    # Mensagem de conclusão
    completion_label = tk.Label(loanding, text="Processo Concluído!", font=("Arial", 14, "bold"), fg="white", bg="#191919")
    completion_label.pack(pady=20)

    # Botão de fechar
    close_button = tk.Button(loanding, text="Fechar", command=loanding.quit, font=("Arial", 12), bg="#4CAF50", fg="white")
    close_button.pack(pady=10)

# Função principal para geração do arquivo Excel
def generate_excel():
    start_button.config(state=tk.DISABLED)  # Desabilitar botão ao iniciar o processo
    # Atualizar progresso: Iniciando
    update_progresso(0, "Iniciando...")

    # Importando dados no estoque

    cominho_arquivos = r"\\fs010J\grpctlest$\00° Indicadores setor Controle de Estoque 2023\QUERY 2024\App.Auto\Alimentação dados"

    list_status = []
    update_progresso(20, "Importando dados...")
    for dados_status in os.listdir(cominho_arquivos):
        if dados_status.startswith('Consulta_Status_de') and dados_status.endswith('.xlsx') or dados_status.endswith(
                '.xls'):
            status = os.path.join(cominho_arquivos, dados_status)

            df_status = pd.read_excel(status, header=None)

            df_status.columns = df_status.iloc[1]

            df_status = df_status.drop([0, 1]).reset_index(drop=True)

            list_status.append(df_status)
    status = pd.concat(list_status, axis=0, ignore_index=True)

    # Atualizar progresso: Importando dados

    update_progresso(30, "Importando dados...")


    status_geral = pd.DataFrame(status)

    update_progresso(40, "Processando dados...")

    data_atual = datetime.now()
    data_hoje = data_atual.date()

    colunas_list = ["Item", "Descplmer", "Peso Unitário", "Unid. Cxa", "Norma Paletização", "Qtde Palete",
                    "Qtde Mínima Venda", "Estoque Contab.",
                    "Venda TLMKT", "Curva ABC", "Cobertura", "Estoque Uti", "Total Blog", "Qtd. Saldo Pedido",
                    "Ind. MRT Exp.", "Ind. FRN Exp.", "Ind. Peric."]
    status = status[colunas_list]

    update_progresso(50, "Processando dados...")

    # Importando endereços-geral resuprimento

    list_unit = []
    for dados_unit in os.listdir(cominho_arquivos):
        if dados_unit.startswith('Endere') and dados_unit.endswith('.xlsx') or dados_unit.endswith('.xls'):
            unit = os.path.join(cominho_arquivos, dados_unit)

            df_unit = pd.read_excel(unit)

            list_unit.append(df_unit)
    enderecos_geral = pd.concat(list_unit, ignore_index=True)

    enderecos_geral = pd.DataFrame(enderecos_geral)

    update_progresso(60, "Mesclando informações...")

    colunas_list_resu = ["Item", "Descrição do Item", "End.", "UOM", "Nível Mín.",
                         "Nível Máx.", "Área Sep.", "Fracionado?", "Etq Fisico"]
    enderecos_geral = enderecos_geral[colunas_list_resu]

    update_progresso(62, "Mesclando informações...")

    enderecos_geral1 = enderecos_geral.groupby(["Item"])["Etq Fisico"].sum().reset_index()
    geral = pd.merge(enderecos_geral, enderecos_geral1, on="Item", how="left")

    colunas_list_resu1 = ["Item", "Descrição do Item", "End.", "UOM", "Nível Mín.",
                         "Nível Máx.", "Área Sep.", "Fracionado?", "Etq Fisico_y"]

    geral = geral[colunas_list_resu1]

    geral["Etq Fisico"] = geral["Etq Fisico_y"]

    geral["Item"] = geral["Item"].astype(int)
    status["Item"] = status["Item"].astype(int)

    update_progresso(63, "Mesclando informações...")

    query_base = pd.merge(geral, status, left_on="Item", right_on="Item", how="left")
    query_base = pd.DataFrame(query_base)

    update_progresso(64, "Mesclando informações...")

    query_base["MOD"] = query_base["End."].str[0]
    query_base["RUA"] = query_base["End."].str[1:4]
    query_base["PDO"] = query_base["End."].str[5:8]
    query_base["A"] = query_base["End."].str[9]

    update_progresso(65, "Mesclando informações...")


    query_base["Cobertura"] = query_base["Cobertura"].str.replace(',', '.').astype(float)

    update_progresso(68, "Mesclando informações...")

    query_base["RUA"] = query_base["RUA"].astype(int)
    query_base["PDO"] = query_base["PDO"].astype(int)
    query_base["A"] = query_base["A"].astype(int)


    update_progresso(70, "Verificando informações...")

    query_base = query_base.rename(columns={
        "Unid. Cxa": "QT CX",
        "Qtde Palete": "QT PL",
        "Qtde Mínima Venda": "QT MIN",
        "Etq Fisico": "ESTOQ. FIS",
        "Estoque Contab.": "ESTOQ. CON.",
        "Venda TLMKT": "TLMKT",
        "Estoque Uti": "ES. ÚTIL",
        "Total Blog": "BLOQ",
        "Qtd. Saldo Pedido": "SALDO P",
        "Ind. Peric.": "PRL",
        "Nível Mín.": "MIN",
        "Nível Máx.": "MAX",
        "Área Sep.": "Área",
        "Norma Paletização": "NORMA",
        "Curva ABC": "CURVA",
        "Ind. MRT Exp.": "EXP",
        "Ind. FRN Exp.": "EXP ",
        "Fracionado?": "FRAC"
    })

    update_progresso(73, "Verificando informações...")

    colunas_list_query = ["Item", "Descrição do Item", "MOD", "RUA", "PDO", "A", "UOM", "Peso Unitário", "QT CX",
                          "NORMA", "QT PL", "QT MIN", "ESTOQ. FIS", "ESTOQ. CON.",
                          "TLMKT", "CURVA", "Cobertura", "ES. ÚTIL", "BLOQ",
                          "SALDO P", "EXP", "EXP ", "PRL", "MIN",
                          "MAX", "Área", "FRAC"]

    query_base = query_base[colunas_list_query]

    update_progresso(75, "Verificando informações...")

    def converter(x):
        try:
            return int(float(x))
        except (ValueError, TypeError):
            return x
    update_progresso(76, "Verificando informações...")

    status = status.apply(lambda col: col.map(converter))

    update_progresso(78, "Verificando informações...")

    status_geral = status_geral.apply(lambda col: col.map(converter))

    update_progresso(79, "Verificando informações...")

    query_base = query_base.apply(lambda col: col.map(converter))

    update_progresso(80, "Convertendo e organizando...")

    diretorio = pd.read_excel("Diretorio Query.xlsx")

    diretorio_ = diretorio.loc[diretorio["index"] == 1, "Diretorio Query"]
    diretorio_ = str(diretorio_.iloc[0])

    update_progresso(85, "Convertendo e organizando...")

    query_base["PESO"] = query_base["Peso Unitário"].str.replace(',', '.').astype(float)

    colunas_list_query = ["Item", "Descrição do Item", "MOD", "RUA", "PDO", "A", "UOM", "PESO", "QT CX",
                          "NORMA", "QT PL", "QT MIN", "ESTOQ. FIS", "ESTOQ. CON.",
                          "TLMKT", "CURVA", "Cobertura", "ES. ÚTIL", "BLOQ",
                          "SALDO P", "EXP", "EXP ", "PRL", "MIN",
                          "MAX", "Área", "FRAC"]

    query_base = query_base[colunas_list_query]

    query_ressu = query_base

    query_ressu["Ressuprir?"] = query_ressu.apply(
            lambda x: 'SIM' if x['ESTOQ. FIS'] <= x['QT MIN'] and x['ESTOQ. CON.'] >= x['MIN'] else '',
            axis=1
        )
    query_ressu["Suprir"] = query_ressu.apply(
            lambda x: x['ESTOQ. FIS'] - x['MAX'] if x['ESTOQ. FIS'] <= x['QT MIN'] and x['ESTOQ. CON.'] >= x['MIN'] else '',
            axis=1
        )

    query_base = query_base[colunas_list_query]

    arquivo_base = rf"{diretorio_}\QUERY {data_hoje}.xlsx"
    with pd.ExcelWriter(arquivo_base, engine='openpyxl') as writer:
        query_base.to_excel(writer, sheet_name="Query", index=False)
        status.to_excel(writer, sheet_name="Status", index=False)
        status_geral.to_excel(writer, sheet_name="Status Geral", index=False)
        query_ressu.to_excel(writer, sheet_name="Ressuprimento", index=False)

    update_progresso(87, "Convertendo e organizando...")

    workbook = openpyxl.load_workbook(arquivo_base)
    worksheet = workbook["Query"]

    update_progresso(88, "Exportando arquivo...")

    header_fill = PatternFill(start_color="483D8B", end_color="483D8B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    update_progresso(89, "Exportando arquivo...")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    update_progresso(90, "Exportando arquivo...")

    for col_num in range(1, len(query_base.columns) + 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f"{col_letter}1"]
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = header_font

    update_progresso(91, "Exportando arquivo...")

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    update_progresso(92, "Exportando arquivo...")

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    if isinstance(cell.value, str):
                        max_length = max(max_length, len(cell.value))
                    else:
                        max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column].width = adjusted_width

    update_progresso(93, "Exportando arquivo...")


    worksheet = workbook["Ressuprimento"]

    update_progresso(95, "Exportando arquivo...")

    header_fill = PatternFill(start_color="483D8B", end_color="483D8B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    update_progresso(96, "Exportando arquivo...")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    update_progresso(97, "Exportando arquivo...")

    for col_num in range(1, len(query_base.columns) + 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f"{col_letter}1"]
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = header_font

    update_progresso(98, "Exportando arquivo...")

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    update_progresso(99, "Exportando arquivo...")

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    if isinstance(cell.value, str):
                        max_length = max(max_length, len(cell.value))
                    else:
                        max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(arquivo_base)
    workbook.close()

    update_progresso(100, "Concluído!")

    concluido()

def start_excel_generation():
    excel_thread = threading.Thread(target=generate_excel)
    excel_thread.start()

# Inicia a interface
animacao()
loanding.mainloop()
